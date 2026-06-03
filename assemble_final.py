#!/usr/bin/env python3
"""Assemble the final spreadsheet with participant-only columns.

VR (spoken) rows  -> reconstructed verbatim from LLM-classified participant
                     segment indices in vr_work/out_*.json.
TEXT (chat) rows  -> regex extraction of the "you" segments (validated).
Blank rows        -> left blank.
Text is never altered; only AI / researcher turns are removed.

Adds three columns:
  C  Participant responses (timestamps kept)
  D  Participant text only (no timestamps)
  E  Word count
"""
import glob
import json
import re
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

SRC = "Transcripts.xlsx"
OUT = "Transcripts_participant_only.xlsx"
COL_TS = "Participant responses (timestamps kept)"
COL_TXT = "Participant text only"
COL_WC = "Word count"

# ---- load VR classification ----
segs = json.load(open("vr_work/vr_segments.json"))
sel = {}
for f in sorted(glob.glob("vr_work/out_*.json")):
    sel.update(json.load(open(f)))


def turns_vr(pid):
    """Return list of (time, text) for the participant segments."""
    by = {s[0]: s for s in segs[pid]}
    return [(by[i][1], by[i][3]) for i in sorted(sel[pid]) if i in by]


# ---- TEXT extraction (participant label == "you") ----
seg_re = re.compile(r"(\w+)\s*Today at\s*(\d{1,2}:\d{2}\s*[AaPp][Mm])")


def turns_text(t):
    out, prev = [], 0
    for m in seg_re.finditer(t):
        label = m.group(1)
        time = re.sub(r"\s+", " ", m.group(2)).strip()
        msg = t[prev:m.start()].strip()
        if label.lower() == "you" and msg:
            msg = re.sub(r"https?://\S*fireflies\S*", "", msg, flags=re.I).strip()
            if msg:
                out.append((time, msg))
        prev = m.end()
    return out


def classify(t):
    has_spk = "Speaker 1" in t or "Speaker 2" in t
    if has_spk and "Today at" not in t:
        return "VR"
    if "Today at" in t and not has_spk:
        return "TEXT"
    return "OTHER"


wb = openpyxl.load_workbook(SRC)
ws = wb.active
header = [c.value for c in ws[1]]
tcol = next(i for i, h in enumerate(header, 1) if h and "transcript" in str(h).lower())
idcol = next(i for i, h in enumerate(header, 1) if str(h).strip().lower() == "id")

c_ts = ws.max_column + 1
c_txt = c_ts + 1
c_wc = c_txt + 1
ws.cell(1, c_ts, COL_TS)
ws.cell(1, c_txt, COL_TXT)
ws.cell(1, c_wc, COL_WC)
wrap = Alignment(wrap_text=True, vertical="top")

n = {"VR": 0, "TEXT": 0, "BLANK": 0, "OTHER": 0}
for r in range(2, ws.max_row + 1):
    raw = ws.cell(r, tcol).value
    pid = str(ws.cell(r, idcol).value)
    t = (raw or "").strip()
    ws.cell(r, c_ts).alignment = wrap
    ws.cell(r, c_txt).alignment = wrap
    if not t:
        n["BLANK"] += 1
        continue
    fmt = classify(t)
    n[fmt] += 1
    if fmt == "VR" and pid in sel:
        pairs = turns_vr(pid)
    elif fmt == "TEXT":
        pairs = turns_text(t)
    else:
        pairs = []
    if not pairs:
        continue
    ts_text = "\n".join(f"{tm} {tx}" for tm, tx in pairs)   # with timestamps
    only_text = "\n".join(tx for tm, tx in pairs)           # no timestamps
    wc = len(re.findall(r"\b[\w']+\b", " ".join(tx for tm, tx in pairs)))
    ws.cell(r, c_ts, ts_text)
    ws.cell(r, c_txt, only_text)
    ws.cell(r, c_wc, wc)

# match the original Transcript column width for the two text columns
src_w = ws.column_dimensions[get_column_letter(tcol)].width or 142.5
ws.column_dimensions[get_column_letter(c_ts)].width = src_w
ws.column_dimensions[get_column_letter(c_txt)].width = src_w
ws.column_dimensions[get_column_letter(c_wc)].width = 12

wb.save(OUT)
print("wrote", OUT, "| breakdown:", n)
