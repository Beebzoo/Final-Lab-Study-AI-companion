#!/usr/bin/env python3
"""Full modality x disclosure-depth analysis.

- Adds DD_R1, DD_R2, WordCount, TopicSensitivity to the SPSS file.
- Descriptives + Mann-Whitney tests by modality (VR vs Text).
- Ordinal logistic regression of depth on modality, controlling for
  topic sensitivity and word count.
- Saves figures and a results dict (results.json) for the report.
"""
import json
import numpy as np
import pandas as pd
import pyreadstat
from scipy import stats
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from statsmodels.miscmodels.ordinal_model import OrderedModel

VR, TEXT = 1, 2
TS_MAP = {"None": 0, "Low": 1, "Medium": 2, "High": 3}


def norm(x):
    x = str(x).strip()
    return x.lstrip("0") or "0"


# ---------- assemble analysis frame ----------
df, meta = pyreadstat.read_sav("Final Lab study AI companions.sav")
sc = json.load(open("coding_work/all_scores.json"))
import openpyxl
ws = openpyxl.load_workbook("Transcripts_participant_only.xlsx", read_only=True).active
hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
wc = {}
for row in list(ws.iter_rows(values_only=True))[1:]:
    d = dict(zip(hdr, row))
    wc[norm(d["ID"])] = d["Word count"]

df["k"] = df["ID"].map(norm)
df["DD_R1"] = df["k"].map(lambda k: sc.get(k, {}).get("R1"))
df["DD_R2"] = df["k"].map(lambda k: sc.get(k, {}).get("R2"))
df["WordCount"] = df["k"].map(wc)
df["TopicSensitivity"] = df["k"].map(lambda k: TS_MAP.get(sc.get(k, {}).get("topic_sensitivity")))

# ---------- save back to SPSS (preserve metadata) ----------
col_labels = dict(zip(meta.column_names, meta.column_labels))
col_labels.update({
    "DD_R1": "Disclosure depth, Rubric 1 (0-3)",
    "DD_R2": "Disclosure depth, Rubric 2 (0-6)",
    "WordCount": "Participant word count (no timestamps)",
    "TopicSensitivity": "Topic sensitivity (0 None - 3 High)",
})
vvl = dict(meta.variable_value_labels)
vvl["Condition"] = {1: "VR", 2: "Text"}
vvl["TopicSensitivity"] = {0: "None", 1: "Low", 2: "Medium", 3: "High"}
kw = dict(column_labels=col_labels, variable_value_labels=vvl)
if getattr(meta, "original_variable_types", None):
    kw["variable_format"] = dict(meta.original_variable_types)
pyreadstat.write_sav(df.drop(columns=["k"]), "Final Lab study AI companions.sav", **kw)

# ---------- analysis frame ----------
d = df.dropna(subset=["DD_R1", "Condition", "WordCount"]).copy()
d["modality"] = d["Condition"].map({VR: "VR", TEXT: "Text"})
vr, tx = d[d.Condition == VR], d[d.Condition == TEXT]
R = {"n_vr": len(vr), "n_text": len(tx)}


def mw(a, b):
    a = np.asarray(a, float); b = np.asarray(b, float)
    a = a[~np.isnan(a)]; b = b[~np.isnan(b)]
    u, p = stats.mannwhitneyu(a, b, alternative="two-sided")
    rb = 1 - 2 * u / (len(a) * len(b))   # rank-biserial (VR vs Text)
    return dict(vr_mean=a.mean(), vr_med=float(np.median(a)), vr_sd=a.std(ddof=1),
                tx_mean=b.mean(), tx_med=float(np.median(b)), tx_sd=b.std(ddof=1),
                U=float(u), p=float(p), r=float(rb))


R["R1"] = mw(vr.DD_R1, tx.DD_R1)
R["R2"] = mw(vr.DD_R2, tx.DD_R2)
R["WC"] = mw(vr.WordCount, tx.WordCount)
R["spearman_wc_r2"] = dict(zip(("rho", "p"), [float(x) for x in stats.spearmanr(d.WordCount, d.DD_R2)]))

# ---------- ordinal logistic regression ----------
d["text"] = (d.Condition == TEXT).astype(float)
d["wc_z"] = (d.WordCount - d.WordCount.mean()) / d.WordCount.std()
d["ts"] = d.TopicSensitivity.fillna(0).astype(float)


def ordreg(y):
    X = d[["text", "ts", "wc_z"]]
    m = OrderedModel(d[y], X, distr="logit").fit(method="bfgs", disp=False)
    out = {}
    for term in ["text", "ts", "wc_z"]:
        out[term] = dict(coef=float(m.params[term]), OR=float(np.exp(m.params[term])),
                         p=float(m.pvalues[term]),
                         ci=[float(np.exp(m.conf_int().loc[term, 0])),
                             float(np.exp(m.conf_int().loc[term, 1]))])
    return out


R["ord_R1"] = ordreg("DD_R1")
R["ord_R2"] = ordreg("DD_R2")
json.dump(R, open("results.json", "w"), indent=2)
print(json.dumps(R, indent=2))

# ============ FIGURES ============
plt.rcParams.update({"font.size": 11, "axes.spines.top": False, "axes.spines.right": False})
C = {"VR": "#4C72B0", "Text": "#DD8452"}

# Fig 1: R1 distribution by modality (grouped %)
fig, ax = plt.subplots(figsize=(7, 4.2))
ct = pd.crosstab(d.modality, d.DD_R1, normalize="index") * 100
x = np.arange(4); w = 0.38
ax.bar(x - w / 2, ct.loc["VR"], w, label="VR", color=C["VR"])
ax.bar(x + w / 2, ct.loc["Text"], w, label="Text", color=C["Text"])
ax.set_xticks(x); ax.set_xticklabels(["0\nNone", "1\nSurface", "2\nModerate", "3\nDeep"])
ax.set_ylabel("% of participants"); ax.set_xlabel("Disclosure depth (Rubric 1)")
ax.set_title("Disclosure depth by modality"); ax.legend(title="Modality")
fig.tight_layout(); fig.savefig("fig1_depth_by_modality.png", dpi=150); plt.close(fig)

# Fig 2: R2 box/strip by modality
fig, ax = plt.subplots(figsize=(6, 4.2))
data = [tx.DD_R2.values, vr.DD_R2.values]
bp = ax.boxplot(data, vert=True, widths=0.55, patch_artist=True, labels=["Text", "VR"])
for patch, k in zip(bp["boxes"], ["Text", "VR"]):
    patch.set_facecolor(C[k]); patch.set_alpha(0.55)
for i, (vals, k) in enumerate(zip(data, ["Text", "VR"]), 1):
    ax.scatter(np.random.normal(i, 0.06, len(vals)), vals, s=14, color=C[k], alpha=0.6, zorder=3)
ax.set_ylabel("Disclosure depth (Rubric 2, 0-6)")
ax.set_title("Finer-grained disclosure depth by modality")
fig.tight_layout(); fig.savefig("fig2_R2_box.png", dpi=150); plt.close(fig)

# Fig 3: word count vs depth (length is not depth)
fig, ax = plt.subplots(figsize=(7, 4.2))
for k, g in d.groupby("modality"):
    j = g.DD_R2 + np.random.normal(0, 0.08, len(g))
    ax.scatter(g.WordCount, j, s=22, alpha=0.6, color=C[k], label=k)
ax.set_xlabel("Participant word count"); ax.set_ylabel("Disclosure depth (Rubric 2)")
ax.set_title("More words ≠ deeper disclosure"); ax.legend(title="Modality")
fig.tight_layout(); fig.savefig("fig3_wordcount_vs_depth.png", dpi=150); plt.close(fig)

# Fig 4: mean word count by modality
fig, ax = plt.subplots(figsize=(5, 4.2))
means = [vr.WordCount.mean(), tx.WordCount.mean()]
sems = [vr.WordCount.sem(), tx.WordCount.sem()]
ax.bar(["VR", "Text"], means, yerr=sems, capsize=6, color=[C["VR"], C["Text"]], alpha=0.85)
ax.set_ylabel("Mean word count"); ax.set_title("VR produces far more words")
fig.tight_layout(); fig.savefig("fig4_wordcount.png", dpi=150); plt.close(fig)
print("\nfigures + results.json + updated SPSS written")
