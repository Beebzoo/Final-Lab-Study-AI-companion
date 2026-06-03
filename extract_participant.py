#!/usr/bin/env python3
"""
Add a column to Transcripts.xlsx containing ONLY the participant's responses
(AI removed), with timestamps preserved.

Two transcript formats are present:

  TEXT chat (P2): messages are written as "<text> <SpeakerLabel>Today at <time>".
      The participant's label is literally "you"; the AI's label is its name
      (BluePink, DEXLab16, ...). -> keep segments whose label == "you".

  VR / spoken (P1): diarised "<time> Speaker <n> <text>". The AI (Eduardo /
      coach) is identified by its coaching language; the participant is the
      human turns in between, with the researcher's setup ("click start",
      "enjoy it") and debrief ("how did it go", "next door") removed, and the
      post-conversation "results / coach" tail dropped.
"""
import re
import openpyxl

SRC = "Transcripts.xlsx"
OUT = "Transcripts_participant_only.xlsx"
NEWCOL = "Participant responses (timestamps kept)"

# ---- VR helpers ----------------------------------------------------------
AI_MARKERS = [
    r"\bI'?m Eduardo\b", r"\bEduardo\b", r"what'?s on your mind",
    r"how do you feel", r"let'?s dive in", r"glad to be working",
    r"\bUser\b", r"it sounds like", r"that sounds", r"thanks for sharing",
    r"how did (?:that|it) (?:make you feel|go for you)", r"i'?m here (?:to listen|for you)",
    r"welcome to your results", r"your coach", r"tell me more",
    r"what (?:made|do you think)", r"glad to (?:be )?(?:connect|hear)",
]
RESEARCHER_MARKERS = [
    r"click (?:start|begin|on)", r"\benjoy\b", r"door close", r"right outside",
    r"how did it go", r"did you enjoy", r"next door", r"clicked? end",
    r"i'?ll be (?:right )?back", r"join me in the room", r"results next door",
    r"five minutes are starting", r"whenever you feel ready", r"how was it",
    r"did it work", r"is it still talking", r"grab (?:all )?your things",
    r"go to the (?:other )?room", r"by yourself in the room", r"to begin",
    r"you can start", r"whenever you('?re| are) ready", r"in the room next",
]
END_MARKERS = [
    r"welcome to your results", r"your coach", r"transcribed by", r"fireflies",
]
ai_re = re.compile("|".join(AI_MARKERS), re.I)
res_re = re.compile("|".join(RESEARCHER_MARKERS), re.I)
end_re = re.compile("|".join(END_MARKERS), re.I)


def vr_turns(t):
    parts = re.split(r"(\d{1,2}:\d{2})\s+Speaker\s+(\d+)\s+", t)
    return [(parts[k], parts[k + 1], parts[k + 2].strip())
            for k in range(1, len(parts) - 1, 3)]


def extract_vr(t):
    turns = vr_turns(t)
    if not turns:
        return "", "no-turns"
    # score each speaker label for AI-likeness (coaching language)
    scores = {}
    for _, sp, tx in turns:
        scores.setdefault(sp, 0)
        scores[sp] += len(ai_re.findall(tx))
    ai_sp = max(scores, key=scores.get) if scores else None
    flags = []
    if not ai_sp or scores[ai_sp] == 0:
        flags.append("AI-speaker-uncertain")
    # end boundary: first turn that hits a results/coach/fireflies marker
    end_idx = next((idx for idx, (_, _, tx) in enumerate(turns)
                    if end_re.search(tx)), len(turns))
    out = []
    for idx in range(0, end_idx):
        tm, sp, tx = turns[idx]
        if sp == ai_sp:                 # AI line -> skip
            continue
        if res_re.search(tx):           # researcher setup/debrief -> skip
            continue
        if tx:
            out.append(f"{tm} {tx}")
    if len(scores) >= 3:
        flags.append("3+speakers")
    return "  ".join(out), "; ".join(flags)


# ---- TEXT helpers --------------------------------------------------------
seg_re = re.compile(r"(\w+)\s*Today at\s*(\d{1,2}:\d{2}\s*[AaPp][Mm])")


def extract_text(t):
    matches = list(seg_re.finditer(t))
    if not matches:
        return "", "no-segments"
    out = []
    prev_end = 0
    for m in matches:
        label = m.group(1)
        time = re.sub(r"\s+", " ", m.group(2)).strip()
        msg = t[prev_end:m.start()].strip()
        # the label token is glued to msg end; strip a trailing duplicate if present
        if label.lower() == "you" and msg:
            # remove fireflies tail if any
            msg = re.sub(r"https?://\S*fireflies\S*", "", msg, flags=re.I).strip()
            if msg:
                out.append(f"{time} {msg}")
        prev_end = m.end()
    return "  ".join(out), ""


# ---- main ----------------------------------------------------------------
def classify(t):
    has_spk = "Speaker 1" in t or "Speaker 2" in t
    has_today = "Today at" in t
    if has_spk and not has_today:
        return "VR"
    if has_today and not has_spk:
        return "TEXT"
    return "OTHER"


wb = openpyxl.load_workbook(SRC)
ws = wb.active
# locate columns
header = [c.value for c in ws[1]]
tcol = next(i for i, h in enumerate(header, 1) if h and "transcript" in str(h).lower())
idcol = next(i for i, h in enumerate(header, 1) if str(h).strip().lower() == "id")
new_i = ws.max_column + 1
flag_i = new_i + 1
ws.cell(1, new_i, NEWCOL)
ws.cell(1, flag_i, "extraction_note")

samples = {}
counts = {"VR": 0, "TEXT": 0, "OTHER": 0, "BLANK": 0}
flagged = []
for r in range(2, ws.max_row + 1):
    raw = ws.cell(r, tcol).value
    pid = ws.cell(r, idcol).value
    t = (raw or "").strip()
    if not t:
        counts["BLANK"] += 1
        ws.cell(r, flag_i, "blank-source")
        continue
    fmt = classify(t)
    counts[fmt] += 1
    if fmt == "VR":
        out, flag = extract_vr(t)
    elif fmt == "TEXT":
        out, flag = extract_text(t)
    else:
        out, flag = "", "unknown-format"
    ws.cell(r, new_i, out)
    ws.cell(r, flag_i, flag)
    if flag:
        flagged.append((pid, fmt, flag))
    samples[str(pid)] = (fmt, out)

wb.save(OUT)
print("counts:", counts)
print("flagged rows needing review:", len(flagged))
for pid, fmt, fl in flagged:
    print(f"   ID {pid} [{fmt}]: {fl}")

print("\n========== SAMPLE OUTPUTS (your examples) ==========")
for pid in ["14140", "14965"]:  # VR examples to eyeball
    if pid in samples:
        print(f"\n--- VR ID {pid} ---\n{samples[pid][1][:900]}")
# find the DEXLab16 (boss) and BluePink (deadline) text examples
for r in range(2, ws.max_row + 1):
    raw = (ws.cell(r, tcol).value or "")
    if "DEXLab16" in raw and "raising my salary" in raw:
        print(f"\n--- TEXT DEXLab16 ID {ws.cell(r,idcol).value} ---\n{ws.cell(r,new_i).value[:900]}")
    if "BluePink" in raw and "missed a deadline" in raw.lower():
        print(f"\n--- TEXT BluePink ID {ws.cell(r,idcol).value} ---\n{ws.cell(r,new_i).value[:900]}")
